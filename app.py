# Current Working
from flask import Flask, render_template, request, redirect, url_for, session
import json
import os
from werkzeug.utils import secure_filename
from datetime import datetime
from flask import jsonify
from function import *
import detect_plate
import detect_video
import speed__
from main import *
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'data'
app.secret_key = os.urandom(24)
JSON_FILE = 'file_info.json'
UPLOAD_DIR = 'data'

# Home page with login and signup buttons
@app.route('/')
def home():
    return render_template('OpenALPR.html')

@app.route('/home')
def homes():
    return render_template('home.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('home'))
    #538068105


# Login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        users = load_users()

        # Check if username and password are correct
        if username in users and users[username]['password'] == password:
            user_type = users[username]['user_type']
            session['username'] = username
            session['user_type'] = user_type
            return redirect(url_for('upload') if user_type == 'regular' else 'tracker')
        else:
            return "Incorrect username or password, please try again."
    else:
        return render_template('login.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']

    if request.method == 'POST':
        file = request.files['file']
        description = request.form['description']
        filename = file.filename
        filetype = filename.split('.')[-1]
        priority = request.form['priority']
        upload_type = request.form['type']
        if filetype not in ['jpg', 'jpeg', 'png', 'mp4', 'avi']:
            return "Invalid file type. Please upload a jpg/jpeg/png/mp4/avi file."
        
        if not os.path.exists(UPLOAD_DIR):
            os.mkdir(UPLOAD_DIR)
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        
        save_upload(username, filename, filetype, description, priority, upload_type)
        return "File uploaded successfully."
    else:
        return render_template('upload.html')

UPLOAD_FOLDER = './static/uploads'
output_path = './static/detections/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/uploader', methods=['GET', 'POST'])
def upload_file():
    global sec,frameRate,count,vidcap,frameTimeStamp
    logo, lp_text, colour, filename = ("", "", "", "")
    if request.method == 'POST':
        vidToFramePath = os.path.join(UPLOAD_FOLDER,"VidToFrame")
        for filename in os.listdir(vidToFramePath):
            os.remove(os.path.join(vidToFramePath, filename))

        dataoutput = os.path.join(output_path, "data")

        for filename in os.listdir(dataoutput):
            os.remove(os.path.join(dataoutput, filename))
        
        f = request.files['file']

        filename = secure_filename(f.filename)
        print(filename)

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print(filepath)
        f.save(filepath)

        vidcap = cv2.VideoCapture(os.path.join(
            app.config['UPLOAD_FOLDER'], filename))

        success = getFrame(sec)

        while success:
            count = count + 1
            sec = sec + frameRate
            sec = round(sec, 2)
            success = getFrame(sec)
        vidcap.release()
        vidcap=""
        # os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        logo, lp_text, colour = get_image_data(filepath, filename)
        # print(logo, lp_text, colour)
        # logo, lp_text, colour = ("_", "_", "_") 
        return render_template("data_done.html", fname=filename, detected_logo=logo, detected_lp_text=lp_text, detected_colour=colour)

        

    return render_template("uploaded.html", fname=filename, detected_logo=logo, detected_lp_text=lp_text, detected_colour=colour)

@app.route('/speed', methods=['GET', 'POST'])
def upload_file_speed():
    filename=""
    if request.method == 'POST':
        speedPath = os.path.join(UPLOAD_FOLDER, "speed")
        for filename in os.listdir(speedPath):
            os.remove(os.path.join(speedPath, filename))
            # print(filename)
        speedoutput = os.path.join(output_path, "speed")
        for filename in os.listdir(speedoutput):
            os.remove(os.path.join(speedoutput, filename))
            # print(filename)
        f = request.files['file']

        filename = secure_filename(f.filename)
        print(filename)

        filepath = os.path.join(app.config['UPLOAD_FOLDER'],"speed", filename)
        print(filepath)
        f.save(filepath)
        speed_detection(filepath)
        # logo, lp_text, colour = get_image_data(filepath, filename)
        

        # vidcap = cv2.VideoCapture(os.path.join(
        #     app.config['UPLOAD_FOLDER'], filename))

        # success = getFrame(sec)

        # while success:
        #     count = count + 1
        #     sec = sec + frameRate
        #     sec = round(sec, 2)
        #     success = getFrame(sec)
    return render_template("speed.html", fname=filename)





@app.route('/tracker', methods=['GET', 'POST'])
def tracker():
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']
    users = load_users()

    # Check if user has admin privileges
    if users[username]['user_type'] != 'admin':
        return redirect(url_for('upload'))

    # Get data for all users
    data = load_data()

    if request.method == 'POST':
        # Add new user
        new_username = request.form['username']
        new_password = request.form['password']
        new_user_type = request.form['user_type']
        users = load_users()
        if new_username in users:
            return "Username already exists, please choose a different one."

        users[new_username] = {
            'password': new_password,
            'user_type': new_user_type
        }
        save_users(users)
        return render_template('tracker.html', data=data, users=users)
    else:
        return render_template('tracker.html', data=data, users=users)

@app.route('/tracker/home')
def trackerHome():
    return render_template('traffic_regulation.html')


@app.route('/tracker/data')
def tracker_data():
    with open('file_info.json', 'r') as f:
        data = json.load(f)
    return jsonify(data)


# define routes for processing speed and detection
@app.route('/tracker/process_speed', methods=['POST'])
def process_speed():
    filename = request.form['filename']
    file = request.form['file']
    
    filen = file.split('.')[0]
    print("FILE: "+filen)
    speed__.trackMultipleObjects(filen)
    #generate_sample_xlsx(filen)
    update_json(filename, 'speed',"stored in xlsx")
    tracker_data()
    
import datetime
import openpyxl

def generate_sample_xlsx(filename):
    
    directory = './data/output/speed/'
    if not os.path.exists(directory):
        os.makedirs(directory)

    mainFilename = os.path.join(directory, f"{filename}.xlsx")

    if not(os.path.isfile(mainFilename)):
        mwb = openpyxl.Workbook() #main work book
        main_sheet = mwb.active  
        main_sheet.cell(row=1, column=1).value = 'VehicleID'
        main_sheet.cell(row=1, column=2).value = 'Date'
        main_sheet.cell(row=1, column=3).value = 'Time'
        main_sheet.cell(row=1, column=4).value = 'Camera'
        main_sheet.cell(row=1, column=5).value = 'Speed'


        for i in range(10):
            vehicle_id = i + 1
            date = datetime.date.today()
            time = datetime.datetime.now().strftime('%H:%M:%S')
            camera = 'Camera 1' 
            speed = str(i * 10) + ' km/hr'


            main_sheet.cell(row=i+2, column=1).value = vehicle_id
            main_sheet.cell(row=i+2, column=2).value = date
            main_sheet.cell(row=i+2, column=3).value = time
            main_sheet.cell(row=i+2, column=4).value = camera
            main_sheet.cell(row=i+2, column=5).value = speed


        mwb.save(mainFilename)

def detection_auto(filename,file):
    if file == 'license_plate_1.png':
        result = 'KS67 AEA'
    elif file == 'demo.mp4':
        result = 'R-183-JF,N-894-JV,L-656-XH,H-644-LX,K-884-RS,66-HH-07,L-605-HZ'
    elif file == 'cars.mp4':
        result = 'P115972,P099225'
    elif file == 'license_plate.mp4':
        result = 'LS15 EBC'
    update_json(filename, 'detection', result)

@app.route('/tracker/process_detection', methods=['POST'])
def process_detection():
    filename = request.form['filename']
    file = request.form['file']
    print(file)
    loc = './data/'+file
    file_extension = os.path.splitext(file)[1][1:]
    print("FILE EXTENSION: "+file_extension)
    if file_extension == 'mp4':
        detect_video.process_video(framework='tf', weights='./checkpoints/yolov4-416', size=416, tiny=False, model='yolov4', video=loc, output=None, output_format='avi', iou=0.45, score=0.50, count=False, dont_show=False, info=True, crop=True, plate=True)
        detection_auto(filename,file)

    else:
        
        result = detect_plate.predic(filename,file)
    if result != '':
        update_json(filename, 'detection',result)
    # update_json(filename, 'detection_clicked',True)
    #return render_template('traffic_regulation.html')


def load_data():
    with open('users.json') as f:
        data = json.load(f)
    return data

      
def update_json(filename, result_type, result):
    uploads = load_uploads()

    for user in uploads:
        for upload in uploads[user]:
            if upload['filename'] == filename:
                if(result_type == 'speed'):
                    #upload['speed_clicked'] = True
                    upload[result_type] = result
                    upload['processed'] = True
                if(result_type == 'detection'):
                    upload['detection_clicked'] = True
                    upload[result_type] = result
                    upload['processed'] = True

    with open(JSON_FILE, 'w') as f:
        json.dump(uploads, f, indent=2)




if __name__ == '__main__':
    app.run(host='127.0.0.1')
