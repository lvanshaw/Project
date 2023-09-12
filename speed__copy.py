import cv2
import dlib
import time
import threading
import math
import datetime
import os
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from skimage import util
from skimage.io import imread
from skimage.filters import threshold_otsu
from skimage import measure
from skimage.measure import regionprops
import matplotlib.patches as patches
import sklearn
import joblib


SNo = int(1)
counter = 0
sheet = {}
now = datetime.datetime.now()
csv =""
if not os.path.exists('./data/output/'):
    os.makedirs('./data/output/')
fileLoc = ""

def create(file):
    mainFilename = file+'.xlsx'
    fileLoc = './data/'+file
    if not(os.path.isfile('./data/output/'+mainFilename)):
        mwb = openpyxl.Workbook() #main work book
        #main_sheet = mwb.create_sheet('main_sheet')
        #main_sheet = mwb.worksheets
        main_sheet=mwb.active  
        #main_sheet = main_sheet.title
        main_sheet.cell(row=1, column=1).value = 'VehicleID'
        main_sheet.cell(row=1, column=2).value = 'Vehicle_Number'


        sheet[0] = mwb.create_sheet('sheet%d' %counter)
        # ws[counter] = mwb.worksheets[counter]
        sheet[0] = mwb.active
        a = 'Daily Vehicles'
        day = now.day
        b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
        sheet[0].title = b
        sheet[0].cell(row=1, column=1).value = 'VehicleID'
        sheet[0].cell(row=1, column=2).value = 'Date'
        sheet[0].cell(row=1, column=3).value = 'Time'
        sheet[0].cell(row=1, column=4).value = 'Camera'
        sheet[0].cell(row=1, column=5).value = 'Speed'
        sheet[0].cell(row=1, column=6).value = 'Number Plate'

        mwb.save('./data/output/'+mainFilename)
        csv = './data/output/'+mainFilename
        trackMultipleObjects()
        
print("FileLocation: "+fileLoc)
carCascade = cv2.CascadeClassifier('myhaar.xml')
video = cv2.VideoCapture(fileLoc)

WIDTH = 1280
HEIGHT = 720

######

def detection(image_1):
    #print(img)
    height = 900
    width = 700
    img = cv2.cvtColor(image_1, cv2.COLOR_RGB2GRAY)
    car_image = cv2.resize(img, (width,height))
    img_1 = cv2.resize(img, (width,height))

    gray_car_image = car_image * 255
    #fig, (ax1, ax2) = plt.subplots(1, 2)
    #ax1.imshow(gray_car_image, cmap="gray")
    threshold_value = threshold_otsu(gray_car_image)
    binary_car_image = gray_car_image > threshold_value

    label_image = measure.label(binary_car_image)

    plate_dimensions = (0.06*label_image.shape[0], 0.13*label_image.shape[0], 0.085*label_image.shape[1], 0.2*label_image.shape[1])
    plate_dimensions2 = (0.06*label_image.shape[0], 0.18*label_image.shape[0], 0.15*label_image.shape[1], 0.25*label_image.shape[1])
    min_height, max_height, min_width, max_width = plate_dimensions
    plate_objects_cordinates = []
    plate_like_objects = []

    flag =0
    go = 0
    
    for region in regionprops(label_image):

        
        if (region.area < 500):
            continue

        print(type(region.bbox))
        print(region.bbox)
        
        min_row, min_col, max_row, max_col = region.bbox

        region_height = max_row - min_row
        region_width = max_col - min_col
        if (region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):
            flag = 1
            go = go + 1
            print(go)

            plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
            plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
            rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor="red", linewidth=2, fill=False)

    
    if(flag == 1):
        print(plate_like_objects[0])
        #plt.show()
        #print('flag=1')


    if(flag == 0):
        min_height, max_height, min_width, max_width = plate_dimensions2
        plate_objects_cordinates = []
        plate_like_objects = []

        for region in regionprops(label_image):
            if (region.area < 150): ## 50 -> 30
                #if the region is so small then it's likely not a license plate
                continue
                # the bounding box coordinates
            min_row, min_col, max_row, max_col = region.bbox

            region_height = max_row - min_row
            region_width = max_col - min_col

    
            # ensuring that the region identified satisfies the condition of a typical license plate
            if (region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):
                # print("hello")
                plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
                plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
                rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor="red", linewidth=2, fill=False)

    if (len(plate_like_objects)!=0):
        print(len(plate_like_objects))
    
        (r1, c1, r2, c2)=plate_objects_cordinates[len(plate_like_objects)-1]
    
        cropped_image_1 = gray_car_image[r1:r2, c1:c2]
        img_2 = img_1[r1:r2, c1:c2]
        img_3 = cv2.medianBlur(img_2,5)
    #print(cropped_image)
    #grayscale = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
        ret,th_1 = cv2.threshold(img_3,155,255,cv2.THRESH_BINARY)
    
        th_2 = cv2.adaptiveThreshold(img_3,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,11,2)# this one is better after all the thresholdings
    
    #th_3 = cv2.adaptiveThreshold(img_3,255,cv2.ADAPTIVE_THRESH_MEAN_C,\
    #            cv2.THRESH_BINARY,11,2)
    
        not_th_2 = cv2.bitwise_not(th_2)
    
    
        width_plate = 90
        height_plate = 60
    
        license_plate = cv2.resize(not_th_2, (width_plate,height_plate))
        labelled_plate = measure.label(license_plate)
    #print(license_plate.shape[0]) #height of the plate
        
        character_dimensions = (0.1875*license_plate.shape[0], 0.480*license_plate.shape[0], 0.05*license_plate.shape[1], 0.40*license_plate.shape[1])
        min_height, max_height, min_width, max_width = character_dimensions #from resized license plate
    #print(character_dimensions)
    
        characters = []
        counter=0
        column_list = []
    
    #fig, (ax1) = plt.subplots(1)
    #ax1.imshow(license_plate, cmap="gray")
        
        for regions in regionprops(labelled_plate):
            y0, x0, y1, x1 = regions.bbox
            region_height = y1 - y0
            region_width = x1 - x0
        
            if (region_height > min_height and region_height < max_height and region_width > min_width and region_width < max_width):
                roi = license_plate[y0:y1, x0:x1]
    
            #print('')
            #print(roi)
            #print('')
            
            # draw a red bordered rectangle over the character.
                rect_border = patches.Rectangle((x0, y0), x1 - x0, y1 - y0, edgecolor="red", linewidth=2, fill=False)
            #print(rect_border)
            
            #ax1.add_patch(rect_border)
    
            # resize the characters to 20X20 and then append each character into the characters list
                resized_char = cv2.resize(roi, (20, 20))
                characters.append(resized_char)
            #print('')
            #print(resized_char)
            
            # this is just to keep track of the arrangement of the characters
                column_list.append(x0)
        filename = 'finalized_model.sav' # write the full directory path and also put double slashes if required!
    
        current_dir = os.path.dirname(os.path.realpath(filename))
    
        model_dir = os.path.join(current_dir, 'finalized_model.sav')
    
        model = joblib.load(model_dir)
    
    #print('Model loaded. Predicting characters of number plate')
    
        classification_result = []
        
        for each_character in characters:
        # converts it to a 1D array
            each_character = each_character.reshape(1, -1);
        #print(each_character)
            result = model.predict(each_character)
            classification_result.append(result)
    
    #print('Classification result')
    #print(classification_result)
    
        plate_string = ''
    
        for eachPredict in classification_result:
            plate_string += eachPredict[0]


        column_list_copy = column_list[:]
        column_list.sort()
        rightplate_string = ''

        for each in column_list:
            rightplate_string += plate_string[column_list_copy.index(each)]

    #print('License plate')
    #print(rightplate_string)

    else:
        plate_string = 'not able to detect'

    return plate_string



######

def estimateSpeed(location1, location2, fps):
    d_pixels = math.sqrt(math.pow(location2[0] - location1[0], 2) + math.pow(location2[1] - location1[1], 2))
    # ppm = location2[2] / carWidht
    ppm = 10                          #$$$$
    d_meters = d_pixels / ppm
    #print("d_pixels=" + str(d_pixels), "d_meters=" + str(d_meters))
    #fps = 10#$$$$
    print(fps)
    speed = d_meters * 10 * 3.6 #3600/1000
    return speed
	

def trackMultipleObjects():
    rectangleColor = (0, 255, 0)
    frameCounter = 0
    currentCarID = 0
    fps = 1
    cropped_image = {}
    LPN = {}
    carTracker = {}
    carNumbers = {}
    carLocation1 = {}
    carLocation2 = {}
    speed = [None] * 1000
    counter = 0
    a = 0
    myuse = []
    	
    # Write output to video file
    out = cv2.VideoWriter('outpy.avi',cv2.VideoWriter_fourcc('M','J','P','G'), 10, (WIDTH,HEIGHT))


    while True:

        now = datetime.datetime.now()
        if (now.hour == 0 & now.minute == 0 & now.second == 0):
            a = 1
            counter = counter + 1
            filename = csv
            mwb = load_workbook(filename)
            sheet[counter] = mwb.create_sheet('sheet%d' %counter)
            # ws[counter] = mwb.worksheets[counter]
            sheet[counter] = mwb.active
            a = 'Daily Vehicles'
            day = now.day
            b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
            sheet[counter].title = b
            sheet[counter].cell(row=1, column=1).value = 'VehicleID'
            sheet[counter].cell(row=1, column=2).value = 'Date'
            sheet[counter].cell(row=1, column=3).value = 'Time'
            sheet[counter].cell(row=1, column=4).value = 'Camera'
            sheet[counter].cell(row=1, column=5).value = 'Speed'
            sheet[counter].cell(row=1, column=6).value = 'Number Plate'
            mwb.save(filename)
                
        start_time = time.time()
        rc, image_2 = video.read()
        if type(image_2) == type(None):
            break
		
        image_1 = cv2.resize(image_2, (WIDTH, HEIGHT))
        resultImage = image_1.copy()
        height = image_1.shape[0]
        width = image_1.shape[1]
        image = image_1[0:height, 0:int(width/2)]
		
        frameCounter = frameCounter + 1
		
        carIDtoDelete = []

        for carID in carTracker.keys():
            trackingQuality = carTracker[carID].update(image)
			
            if trackingQuality < 7: #7
                carIDtoDelete.append(carID)
				
        for carID in carIDtoDelete:

            carTracker.pop(carID, None)
            carLocation1.pop(carID, None)
            carLocation2.pop(carID, None)
            cropped_image.pop(carID,None)
            LPN.pop(carID, None)
		
        if not (frameCounter % 10): #it goes inside only when frameCounter is a multiple of 10
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            cars = carCascade.detectMultiScale(gray, 1.1, 13, 18, (24, 24))

            cars = list(cars)

            #print(cars)
            
            for i in range (len(cars)):
                #print(cars, 'HAKUNA MATATA')
                if (cars[i][0] < 195) or (cars[i][2] < 58) or (cars[i][3] < 58) or ((cars[i][1] > 40) and (cars[i][2] < 68)) or ((cars[i][1] > 268) and (cars[i][2] < 250)) or ((cars[i][1] > 268) and (cars[i][3] < 250)) :
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 40) and (cars[i][2] < 70):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 110) and (cars[i][2] < 120):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 268) and (cars[i][2] < 250):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for (_x, _y, _w, _h) in cars:
                x = int(_x)
                y = int(_y)
                w = int(_w)
                h = int(_h)
			
                x_bar = x + 0.5 * w
                y_bar = y + 0.5 * h
				
                matchCarID = None
                
		
                for carID in carTracker.keys():
                    trackedPosition = carTracker[carID].get_position()
		    
                    t_x = int(trackedPosition.left())
                    t_y = int(trackedPosition.top())
                    t_w = int(trackedPosition.width())
                    t_h = int(trackedPosition.height())

                    					
                    t_x_bar = t_x + 0.5 * t_w
                    t_y_bar = t_y + 0.5 * t_h
				
                    if ((t_x <= x_bar <= (t_x + t_w)) and (t_y <= y_bar <= (t_y + t_h)) and (x <= t_x_bar <= (x + w)) and (y <= t_y_bar <= (y + h))):
                        matchCarID = carID
				
                if matchCarID is None:
                    print ('Creating new tracker ' + str(currentCarID))
				    
                    tracker = dlib.correlation_tracker()
                    tracker.start_track(image, dlib.rectangle(x, y, x + w, y + h))
					
                    carTracker[currentCarID] = tracker
                    carLocation1[currentCarID] = [x, y, w, h]

                    currentCarID = currentCarID + 1
		
		#cv2.line(resultImage,(0,480),(1280,480),(255,0,0),5)

        myuse = []
        #print(cars)
        #print(carTracker, 'BEFOREEEE')

        ###1
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')
                    if (x_2 < (x_1 + 10)) and ((x_1 + 10) < (x_2 + w_2)) and ((x_1 + w_1 - 10) > x_2) and ((x_1 + w_1 - 10) < (x_2 + w_2)) and ((y_1 + h_1 + 10) > y_2) and ((y_1 + h_1 +10) < (y_2 + h_2)):
                        myuse.append(carID)
                        #print('HAKUNAAAAA')

        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###2
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 + w_1 + 10) > x_2) and ((x_1 + w_1 + 10) < (x_2 + w_2)) and ((y_1 + 10) > y_2) and ((y_1 + 10) < (y_2 + h_2)) and ((y_1 + h_1 - 10) > y_2) and ((y_1 + h_1 - 10) < (y_2 + h_2)) :
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###3
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 + 10) > x_2) and ((x_1 + 10) < (x_2 + w_2)) and ((x_1 + w_1 - 10) > x_2) and ((x_1 + w_1 - 10) < (x_2 + w_2)) and ((y_1 - 10) > y_2) and ((y_1 - 10) < (y_2 + h_2)):
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###4
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 - 10) > x_2) and ((x_1 - 10) < (x_2 + w_2)) and ((y_1 + 10) > y_2) and ((y_1 + 10) < (y_2 + h_2)) and ((y_1 + h_1 - 10) > y_2) and ((y_1 + h_1 - 10) < (y_2 + h_2)) :
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        for carID in carTracker.keys():
            trackedPosition = carTracker[carID].get_position()
					
            t_x = int(trackedPosition.left())
            t_y = int(trackedPosition.top())
            t_w = int(trackedPosition.width())
            t_h = int(trackedPosition.height())
			
            cv2.rectangle(resultImage, (t_x, t_y), (t_x + t_w, t_y + t_h), rectangleColor, 4)

            carLocation2[carID] = [t_x, t_y, t_w, t_h]
		
        end_time = time.time()
		
        if not (end_time == start_time):
            fps = 1.0/(end_time - start_time)

        for i in carLocation1.keys():	
            if frameCounter % 1 == 0: #it will always work
                [x1, y1, w1, h1] = carLocation1[i]
                [x2, y2, w2, h2] = carLocation2[i]
		
		# print 'previous location: ' + str(carLocation1[i]) + ', current location: ' + str(carLocation2[i])
                carLocation1[i] = [x2, y2, w2, h2]
		
		# print 'new previous location: ' + str(carLocation1[i])
                if [x1, y1, w1, h1] != [x2, y2, w2, h2]:
                    if (speed[i] == None or speed[i] == 0): #275.. 285  and y1 >= 250 and y1 <= 260.. this is to give command to the code to only estimate speed when the vehicle is between 275 and 285! 
                        speed[i] = estimateSpeed([x1, y1, w1, h1], [x2, y2, w2, h2], fps)
                        print('1st', x1, y1, w1, h1, ' ', '2nd', x2, y2, w2, h2, ' ', speed[i], ' ', 'fps:', fps)
                        #print('')

			#if y1 > 275 and y1 < 285:
                    if (speed[i] != None and (x2 > x1+5 or x2 < x1-5) and (y2 > y1+5 or y2 < y1-5)): #so that even if the driver opens his seat, the code doesn't detect speed in that
                        cv2.putText(resultImage, str(int(speed[i])) + " km/hr", (int(x1 + w1/2), int(y1-5)),cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 0, 0), 2)
                        print(speed[i], 'YOOOOOOOOOOOO')
                        print(datetime.datetime.now())

                        
                        LPN[i] = detection(cropped_image[i])
                        print(LPN[i])
        
                        

                        a = int(a) + int(1)
                        a = int(a)
                        c = datetime.datetime.now()
                        d = c.strftime('%d') + '/' + c.strftime('%b') + '/' + c.strftime('%Y')
                        e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')
                        

                        filename = csv
                        mwb = load_workbook(filename)
                        sheet[counter] = mwb.active
                        max_row_sheet = sheet[counter].max_row
                        max_row_excel = int(max_row_sheet) + 1
                        sheet[counter].cell(row=max_row_excel, column=1).value = int(max_row_sheet)
                        sheet[counter].cell(row=max_row_excel, column=2).value = d
                        sheet[counter].cell(row=max_row_excel, column=3).value = e
                        sheet[counter].cell(row=max_row_excel, column=4).value = 'Cam1'
                        sheet[counter].cell(row=max_row_excel, column=5).value = speed[i]
                        sheet[counter].cell(row=max_row_excel, column=6).value = LPN[i]

                        mwb.save(filename)
                        
        cv2.imshow('result', resultImage)



        if cv2.waitKey(33) == 27:
            break
	
    cv2.destroyAllWindows()

# if __name__ == '__main__':
    
#     trackMultipleObjects()
